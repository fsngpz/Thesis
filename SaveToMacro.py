from openpyxl import load_workbook
import datetime

def save(sn, st, fn, lp, ln, ss, ls, mn, ll, kpk, mj, dc, ra, rm):
    shift = 0
    start1 = datetime.time(0, 0, 0)
    end1 = datetime.time(8, 0, 0)
    start2 = datetime.time(8, 0, 1)
    end2 = datetime.time(16, 0, 0)
    start3 = datetime.time(16, 0, 1)
    end3 = datetime.time(23, 59, 59)
    dateToday = datetime.date.today()

    if start1 <= datetime.datetime.now().time() <= end1:
        shift = 1
    elif start2 <= datetime.datetime.now().time() <= end2:
        shift = 2
    elif start3 <= datetime.datetime.now().time() <= end3:
        shift = 3
    try:
        if st == 'PASS':
            if lp == "LBO":
                if shift == 1:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 1.xlsm"
                elif shift == 2:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 2.xlsm"
                elif shift == 3:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 3.xlsm"
                wb = load_workbook(path, read_only=False, keep_vba=True)
                ws = wb.worksheets[0]
                ws_tables = []
                lrow = len(ws['A']) + 1
                print(lrow)
                ws[f"A{lrow}"] = dateToday
                ws[f"B{lrow}"] = shift
                ws[f"C{lrow}"] = "IME"
                ws[f"D{lrow}"] = fn
                ws[f"E{lrow}"] = ln
                ws[f"F{lrow}"] = ss
                ws[f"I{lrow}"] = ls
                ws[f"J{lrow}"] = mn
                ws[f"K{lrow}"] = ll
                ws[f"L{lrow}"] = sn

                for table in ws._tables:
                    ws_tables.append(table)

                wb.save(path)
                return "Yeay, success to save the key-in record to Macro Excel file."
            elif lp == "Patrol":
                if shift == 1:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 1.xlsm"
                elif shift == 2:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 2.xlsm"
                elif shift == 3:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 3.xlsm"
                    wb = load_workbook(path, read_only=False, keep_vba=True)
                    ws = wb.worksheets[0]
                    ws_tables = []
                    lrow = len(ws['A']) + 1
                    print(lrow)
                    ws[f"A{lrow}"] = dateToday
                    ws[f"B{lrow}"] = shift
                    ws[f"C{lrow}"] = "PATROL-IME"
                    ws[f"D{lrow}"] = fn
                    ws[f"E{lrow}"] = ln
                    ws[f"F{lrow}"] = ss
                    ws[f"I{lrow}"] = mn
                    ws[f"J{lrow}"] = ll
                    ws[f"K{lrow}"] = sn

                    for table in ws._tables:
                        ws_tables.append(table)

                    wb.save(path)
                    return "Yeay, success to save the key-in record to Macro Excel file."
        elif st == 'FAIL':
            if lp == "LBO":
                if shift == 1:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 1.xlsm"
                elif shift == 2:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 2.xlsm"
                elif shift == 3:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\LBO\LBO Shift 3.xlsm"
                wb = load_workbook(path, read_only=False, keep_vba=True)
                ws = wb.worksheets[0]
                ws_tables = []
                lrow = len(ws['A']) + 1
                print(lrow)
                ws[f"A{lrow}"] = dateToday
                ws[f"B{lrow}"] = shift
                ws[f"C{lrow}"] = "IME"
                ws[f"D{lrow}"] = fn
                ws[f"E{lrow}"] = ln
                ws[f"F{lrow}"] = ss
                ws[f"G{lrow}"] = mj
                ws[f"H{lrow}"] = dc
                ws[f"I{lrow}"] = ls
                ws[f"J{lrow}"] = mn
                ws[f"K{lrow}"] = ll
                ws[f"L{lrow}"] = sn
                ws[f"M{lrow}"] = kpk
                ws[f"Q{lrow}"] = rm
                ws[f"R{lrow}"] = ra

                for table in ws._tables:
                    ws_tables.append(table)

                wb.save(path)
                return "Yeay, success to save the key-in record to Macro Excel file."
            elif lp == "Patrol":
                if shift == 1:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 1.xlsm"
                elif shift == 2:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 2.xlsm"
                elif shift == 3:
                    path = r"C:\Users\User\Documents\President University\Thesis\Inspection Report\PATROL\PATROL Shift 3.xlsm"
                wb = load_workbook(path, read_only=False, keep_vba=True)
                ws = wb.worksheets[0]
                ws_tables = []
                lrow = len(ws['A']) + 1
                print(lrow)
                ws[f"A{lrow}"] = dateToday
                ws[f"B{lrow}"] = shift
                ws[f"C{lrow}"] = "PATROL-IME"
                ws[f"D{lrow}"] = fn
                ws[f"E{lrow}"] = ln
                ws[f"F{lrow}"] = ss
                ws[f"G{lrow}"] = mj
                ws[f"H{lrow}"] = dc
                ws[f"I{lrow}"] = mn
                ws[f"K{lrow}"] = sn
                ws[f"O{lrow}"] = rm
                ws[f"P{lrow}"] = ra


                for table in ws._tables:
                    ws_tables.append(table)

                wb.save(path)
                return "Yeay, success to save the key-in record to Macro Excel file."
    except FileNotFoundError:
        return "Sorry, we cannot found the Macro Excel. Please verify the item's location and try again!"
    except OSError:
        return "Sorry, the Macro Excel file is opened. Please close the file and try again!"
            #st     fn          lp    ln  ss    ls          mn      ll      kpk       dc          ra          rm
# print(save('FAIL', 'Ferdinand', 'Patrol', 2, 32, '151-280',  'WW21', 313687,  219831, "Misshead" ,'PO21',  'BreakPart',  'Watchout'))
